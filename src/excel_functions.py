import os
import pandas as pd
import numpy as np
import re
from dateutil import parser
import datetime
from itertools import pairwise
from openpyxl import load_workbook
from collections import defaultdict


#############################################################################################


reg1 = r"\d{2}-\d{4}"
reg2 = r"\d{2}_\d{4}"
reg3 = r"\d{2}-[a-zA-Z]*_?(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)-\d{4}"
reg4 = r"\d{2}_[a-zA-Z]*_?(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)_\d{4}"
reg5 = r"\d{2}-[a-zA-Z]*-?(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)-\d{4}"
combined_regex = f"({reg1})|({reg2})|({reg3})|({reg4})|({reg5})"
# Correcting wrong patterns:
sub1 = r"-[a-zA-Z]*_"
sub2 = r"-[a-zA-Z]*-"
sub3 = r"_[a-zA-Z]*_"
combined_sub = f"({sub1})|({sub2})|({sub3})"


#############################################################################################


def alphabetic_sequence_to_number(sequence):
    number = 0
    for char in sequence.upper():  # Ensure the sequence is uppercase for simplicity
        if not char.isalpha():  # Skip any non-alphabet characters
            continue
        number = number * 26 + (ord(char) - ord("A") + 1)
    return number


#############################################################################################


def clean_many_dfs(df: pd.DataFrame, threshold_) -> pd.DataFrame:

    num_rows = df.shape[0]
    num_cols = df.shape[1]

    limits = []
    dfs_list = []
    empty_cols = df.columns[df.isna().sum() == len(df)].values
    start = df.iloc[:, :1].columns.tolist()
    end = df.iloc[:, -1:].columns.tolist()
    points = start + empty_cols.tolist() + end

    for one, two in pairwise(points):
        if one == start[0]:
            one = one
        else:
            one = one + 1
        if two == end[0]:
            two = two
        else:
            two = two - 1

        pair = (one, two)
        limits.append(pair)

    for pair in limits:
        temp = df.loc[:, pair[0] : pair[1]]
        threshold = threshold_
        temp.dropna(axis=0, thresh=threshold, inplace=True)
        # Get rid of columns with 90% of NA values
        threshold = 0.1 * num_rows  # 90% threshold for dropping columns
        temp.dropna(axis=1, thresh=threshold, inplace=True)

        dfs_list.append(temp)

    # Checking if all dataframes in the same Excel Sheet have a column giving the date:
    pattern_date = r"date"
    date_count = 0
    for df in dfs_list:
        if (
            df.applymap(lambda x: re.search(pattern_date, str(x), re.IGNORECASE))
            .any()
            .sum()
            > 0
        ):
            date_col = df[
                df.columns[
                    df.applymap(
                        lambda x: re.search(pattern_date, str(x), re.IGNORECASE)
                    ).any(axis=0)
                ]
            ].iloc[:, 0]
            date_count = date_count + 1

    # If not all of them have the date column, then we need to create it.
    if date_count == len(dfs_list):
        return dfs_list
    else:
        for df in dfs_list:
            if (
                df.applymap(lambda x: re.search(pattern_date, str(x), re.IGNORECASE))
                .any()
                .sum()
                == 0
            ):
                df[date_col.name] = date_col
        return dfs_list


#############################################################################################


def infer_series_dtype(series: pd.Series) -> str:
    # Numeric test
    try:
        pd.to_numeric(
            series.dropna().apply(
                lambda x: str(x).replace("-", "0") if "-" in str(x) else str(x)
            )
        )
        # pd.to_numeric(series.dropna())
        return "numeric"
    except:
        pass

    # Datetime test
    try:
        pd.to_datetime(series.dropna())
        return "datetime"
    except:
        pass

    return "string"


#############################################################################################


def retrieve_date(path: str) -> str:
    # Function to extract date strings from paths
    def extract_date_string(path):
        # Adjusted regex to match a broader range of formats
        # This pattern assumes month-year format directly or month-ANYSTRING-year
        match = re.search(combined_regex, path, re.IGNORECASE)
        if match:
            # For formats with ANYSTRING, replace it with a placeholder or remove it
            cleaned_date_str = re.sub(combined_sub, "-", match.group(0))
            return cleaned_date_str
        else:
            return None

    # Function to parse dates
    def parse_date(date_str):
        try:
            return parser.parse(date_str, default=datetime.datetime(1900, 1, 1))
        except ValueError:
            raise ValueError("Unrecognized date format")

    # Extract and parse dates from paths
    date_str = extract_date_string(path)
    date_ok = parse_date(date_str)
    return date_ok


#############################################################################################


def retrieve_year(path: str) -> str:
    # Function to extract date strings from paths
    def extract_date_string(path):
        # Adjusted regex to match a broader range of formats
        # This pattern assumes month-year format directly or month-ANYSTRING-year
        match = re.search(r"\b[0-9]{4}\b", path, re.IGNORECASE)
        if match:
            # For formats with ANYSTRING, replace it with a placeholder or remove it
            cleaned_date_str = re.sub(combined_sub, "-", match.group(0))
            return cleaned_date_str
        else:
            return None

    # Function to parse dates
    def parse_date(date_str):
        try:
            return parser.parse(date_str, default=datetime.datetime(1900, 1, 1))
        except ValueError:
            raise ValueError("Unrecognized date format")

    # Extract and parse dates from paths
    date_str = extract_date_string(path)
    date_ok = parse_date(date_str)
    return date_ok


#############################################################################################


def find_version(input_str: str) -> str:
    # Define the regex pattern
    # The pattern looks for a '-', ' ', or '_' followed by 'v' or 'V', and then one or more digits.
    pattern = r"[- _][vV](\d+)"
    # Perform the search
    match = re.search(pattern, input_str)
    # If a match is found, return it. Otherwise, return None or an appropriate response
    if match:
        return match.group(1)  # Returns the entire match including prefix and 'v'
    else:
        return None


#############################################################################################


def basic_xlsx_clean(df: pd.DataFrame, threshold) -> pd.DataFrame:
    # 1. Get the number of rows in the dataframe
    num_rows = df.shape[0]
    num_cols = df.shape[1]
    # 2. Get rid of rows under the threshold.
    threshold = threshold
    df.dropna(axis=0, thresh=threshold, inplace=True)

    # Check if there are columns that are completely empty.
    if (df.isna().sum() == len(df)).sum() > 0:
        df = clean_many_dfs(df, threshold)
        return df

    else:
        # 3. Get rid of columns with 90% of NA values
        threshold = 0.1 * num_rows  # 90% threshold for dropping columns
        df.dropna(axis=1, thresh=threshold, inplace=True)
        return df


#############################################################################################


def cleaning_by_column_type(df: pd.DataFrame) -> pd.DataFrame:
    num_rows = df.shape[0]
    num_cols = df.shape[1]

    df = df.reset_index(drop=True)
    # 4. Get what type of column is each one by getting the 90%
    indeces_list = []
    df_corrected = pd.DataFrame()
    ## Getting type of column and also what rows should we considers as headers.
    for col in df.columns:
        type = infer_series_dtype(df.loc[num_rows * 0.15 : num_rows * 0.8, col])
        if type == "datetime":
            indeces = (
                pd.to_datetime(df[col], errors="coerce")
                .loc[: num_rows * 0.8][
                    pd.to_datetime(df[col].fillna(0), errors="coerce")
                    .loc[: num_rows * 0.8]
                    .isna()
                ]
                .index
            )
            indeces_list.append(indeces.to_list())
        elif type == "numeric":
            indeces = (
                pd.to_numeric(
                    df[col].apply(
                        lambda x: str(x).replace("-", "0") if "-" in str(x) else str(x)
                    ),
                    errors="coerce",
                )
                .loc[: num_rows * 0.8][
                    pd.to_numeric(
                        df[col]
                        .apply(
                            lambda x: (
                                str(x).replace("-", "0") if "-" in str(x) else str(x)
                            )
                        )
                        .replace("nan", np.nan)
                        .fillna(0),
                        errors="coerce",
                    )
                    .loc[: num_rows * 0.8]
                    .isna()
                ]
                .index
            )
            indeces_list.append(indeces.to_list())
        else:
            continue
    indeces = list(
        range(0, max(indeces_list)[-1] + 1)
    )  ############### The -1 might cause some issues ?? 0 didn't work, sierra_sur.

    # 5. Putting the row values as headers:
    for col in df.columns:
        type = infer_series_dtype(df.loc[num_rows * 0.15 : num_rows * 0.8, col])
        if type == "datetime":
            headers = df.loc[indeces, col].dropna()
            if len(headers) == 0:
                header = "missing"
            else:
                header = "|".join(str(item) for item in headers.dropna().values)
            df_corrected[header] = pd.to_datetime(
                df.loc[max(indeces) + 1 :, col], errors="coerce"
            )
        elif type == "numeric":
            headers = df.loc[indeces, col].dropna()
            if len(headers) == 0:
                header = "missing"
            else:
                header = "|".join(str(item) for item in headers.dropna().values)
            df_corrected[header] = pd.to_numeric(
                df.loc[max(indeces) + 1 :, col]
                .apply(lambda x: str(x).replace("-", "0") if "-" in str(x) else str(x))
                .replace("nan", np.nan)
                .fillna(0),
                errors="coerce",
            )
        else:
            headers = df.loc[indeces, col].dropna()
            if len(headers) == 0:
                header = "missing"
            else:
                header = "|".join(str(item) for item in headers.dropna().values)
            df_corrected[header] = df.loc[max(indeces) + 1 :, col]
    # Dropping data that isn't OK (not from the same format as the vast majority)
    for col in df_corrected.columns:
        df_corrected = df_corrected[~df_corrected[col].isna()]

    return df_corrected


#############################################################################################


def merge_check(df: pd.DataFrame, file_path: str, sheet_name: str) -> pd.DataFrame:

    if any(re.search(r"\|date/time", col, re.IGNORECASE) for col in df.columns) == True:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        ## Merged cells correction:
        wb = load_workbook(filename=file_path)
        merged_cells_ranges = wb[sheet_name].merged_cells.ranges
        merged_cells_info = [str(range) for range in merged_cells_ranges]

        for range in merged_cells_info:
            ranges = range.split(":")
            id1 = ranges[0]
            id2 = ranges[1]
            id1_1 = int(re.search(r"(\d)+", id1).group(0)) - 1
            id1_2 = alphabetic_sequence_to_number(id1) - 1
            id2_1 = int(re.search(r"(\d)+", id2).group(0)) - 1
            id2_2 = alphabetic_sequence_to_number(id2) - 1

            df.iloc[id1_1 : id2_1 + 1, id1_2 : id2_2 + 1] = df.iloc[id1_1, id1_2]

        df = basic_xlsx_clean(df, threshold=2)
        df = cleaning_by_column_type(df)

        return df
    else:
        return df


#############################################################################################


def merged_dfs(df: pd.DataFrame) -> pd.DataFrame:

    if any(re.search(r".*\|(date_time)", col) for col in df.columns):
        date_formatting = re.compile(r".*\|(date_time)")
        for col in df.columns:
            if date_formatting.search(col):
                prev = date_formatting.search(col).group(0)
                now = date_formatting.search(col).group(1)
                df = df.rename(columns={prev: now})

        if np.isin(df["excel_sheet"].unique(), ["TRPT Turb", "AQP Turb", "CJA Turb"])[
            0
        ]:
            df[["variables", "sign"]] = df["variables"].str.split("|", expand=True)
            inverse = df[~df["variables"].str.contains(".Turb", regex=True)].index
            df["sign"] = df["sign"].str.contains(r"(?i)export", regex=True)
            df["sign"] = df["sign"].apply(lambda x: 1 if x == True else -1)
            df.loc[inverse, "sign"] = df.loc[inverse, "sign"] * -1
            df["nominal"] = df["nominal"] * df["sign"]
            df = df.drop(columns=["sign"])
            df = (
                df.groupby([x for x in df.columns if x != "nominal"])["nominal"]
                .sum()
                .reset_index()
            )
        else:
            df[["variables", "sign"]] = df["variables"].str.split("|", expand=True)
            df["sign"] = df["sign"].str.contains(r"(?i)export", regex=True)
            df["nominal"] = df["nominal"] * df["sign"]
            df = df.drop(columns=["sign"])
        return df
    else:
        return df


#############################################################################################


def dropping_totals(df: pd.DataFrame) -> pd.DataFrame:
    reg = re.compile(
        r"\b[tT][oO][-\s]?[tT][aA][-\s]?[lL]\b|\b[tT][oO][tT][aA][lL]\b", re.IGNORECASE
    )
    to_drop = [re.search(reg, col).string for col in df.columns if re.search(reg, col)]
    df = df.drop(columns=to_drop)

    def contains_total(row):
        for cell in row:
            if pd.notnull(cell) and reg.search(str(cell)):
                return True
        return False

    # Apply the function to each row and filter the DataFrame
    filtered_df = df[~df.apply(contains_total, axis=1)]
    return filtered_df


#############################################################################################


def dropping_2(df: pd.DataFrame, fuente: str) -> pd.DataFrame:
    text = fuente.upper()
    reg = re.compile(rf"{text}")
    to_drop = [re.search(reg, col).string for col in df.columns if re.search(reg, col)]
    df = df.drop(columns=to_drop)

    def contains_text(row):
        for cell in row:
            if pd.notnull(cell) and reg.search(str(cell)):
                return True
        return False

    # Apply the function to each row and filter the DataFrame
    filtered_df = df[~df.apply(contains_text, axis=1)]
    return filtered_df


#############################################################################################


def tabular_melting(df: pd.DataFrame) -> pd.DataFrame:
    # Tabulation
    date_cols = list(
        df.select_dtypes(include="datetime")
    )  # All the values that are a date
    string_cols = list(
        df.select_dtypes(include="object")
    )  # All the values that are a string.
    numeric_cols = df.columns[~df.columns.isin(date_cols + string_cols)]
    df = pd.melt(
        df,
        id_vars=date_cols + string_cols,
        value_vars=numeric_cols,
        var_name="variables",
        value_name="nominal",
    )
    return df


#############################################################################################


def excel_sheets_with_revisions(sheets: pd.Series) -> dict:

    # Mapping of month abbreviations to full names
    month_map = {
        "Jan": "January",
        "Feb": "February",
        "Mar": "March",
        "Apr": "April",
        "May": "May",
        "Jun": "June",
        "Jul": "July",
        "Aug": "August",
        "Sep": "September",
        "Sept": "September",
        "Oct": "October",
        "Nov": "November",
        "Dec": "December",
    }

    # Regex to match month names/abbreviations and revisions
    regex = re.compile(
        r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\s*(Rev\s*\d*|rev\s*\d*|\(Rev\s*\d*\)|\(rev\s*\d*\))?",
        re.IGNORECASE,
    )

    # Initialize a dictionary to track the highest revision for each month
    month_revisions = defaultdict(int)
    name = dict()

    for month_str in sheets:
        match = regex.match(month_str.strip())
        if match:
            month_abbr = match.group(1).title()  # Normalize month abbreviation
            month_full = month_map.get(
                month_abbr, month_abbr
            )  # Convert to full month name if abbreviation
            # Extract revision number, defaulting to 0 if not found
            revision_number = 0
            revision_name = month_str
            if match.group(2):
                revision_str = re.findall(r"\d+", match.group(2))
                if revision_str:
                    revision_number = int(revision_str[0])
                    revision_name = month_str

            # Update the dictionary if this revision is higher than what we've previously seen
            if revision_number >= month_revisions[month_full]:
                month_revisions[month_full] = revision_number
                name[month_full] = revision_name

    return name


#############################################################################################


def repeat_merged_cells(
    df: pd.DataFrame, file_path: str, sheet_name: str
) -> pd.DataFrame:

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    ## Merged cells correction:
    wb = load_workbook(filename=file_path)
    merged_cells_ranges = wb[sheet_name].merged_cells.ranges
    merged_cells_info = [str(range) for range in merged_cells_ranges]

    for range in merged_cells_info:
        ranges = range.split(":")
        id1 = ranges[0]
        id2 = ranges[1]
        id1_1 = int(re.search(r"(\d)+", id1).group(0)) - 1
        id1_2 = alphabetic_sequence_to_number(id1) - 1
        id2_1 = int(re.search(r"(\d)+", id2).group(0)) - 1
        id2_2 = alphabetic_sequence_to_number(id2) - 1

        df.iloc[id1_1 : id2_1 + 1, id1_2 : id2_2 + 1] = df.iloc[id1_1, id1_2]

    return df


#############################################################################################


def basic_xlsx_clean_one_table(df: pd.DataFrame, threshold) -> pd.DataFrame:

    # 1. Get the number of rows in the dataframe
    num_rows = df.shape[0]
    num_cols = df.shape[1]
    # 2. Get rid of rows under the threshold.
    threshold = threshold
    df.dropna(axis=0, thresh=threshold, inplace=True)

    # Check if there are columns that are completely empty.
    if (df.isna().sum() == len(df)).sum() == 1:

        part_1 = df.columns[df.isna().sum() == len(df)].values[0] - 0
        part_2 = df.columns[-1] - df.columns[df.isna().sum() == len(df)].values[0]

        if part_1 > part_2:
            df = df.loc[:, : df.columns[df.isna().sum() == len(df)].values[0]]
            # 3. Get rid of columns with 90% of NA values
            threshold = 0.1 * num_rows  # 90% threshold for dropping columns
            df.dropna(axis=1, thresh=threshold, inplace=True)
            return df
        else:
            df = df.loc[:, df.columns[df.isna().sum() == len(df)].values[0] :]
            # 3. Get rid of columns with 90% of NA values
            threshold = 0.1 * num_rows  # 90% threshold for dropping columns
            df.dropna(axis=1, thresh=threshold, inplace=True)
            return df

    else:
        # 3. Get rid of columns with 90% of NA values
        threshold = 0.1 * num_rows  # 90% threshold for dropping columns
        df.dropna(axis=1, thresh=threshold, inplace=True)
        return df


#############################################################################################


def getting_time_series_frame(df: pd.DataFrame) -> pd.DataFrame:

    pattern_dt = re.compile(r"(?i)(?=.*\bdata\b)(?=.*\btime\b)")
    bool_map = df.applymap(lambda x: True if re.search(pattern_dt, str(x)) else False)

    if bool_map.sum().sum() == 0:
        pattern_dt = re.compile(r"(?i)(?=.*\bdate\b)(?=.*\btime\b)")
        bool_map = df.applymap(
            lambda x: True if re.search(pattern_dt, str(x)) else False
        )
        position = bool_map.stack()[bool_map.stack()].index.tolist()[0]
        df = df.loc[position[0] :, position[1] :]
    else:
        position = bool_map.stack()[bool_map.stack()].index.tolist()[0]
        df = df.loc[position[0] :, position[1] :]
    return df


#############################################################################################


def dropping_extras(df: pd.DataFrame, list_to_drop) -> pd.DataFrame:

    def contains_extra(row):
        for cell in row:
            if pd.notnull(cell) and (item in str(cell).lower()):
                return True
        return False

    for item in list_to_drop:
        to_drop = [col for col in df.columns if (item in str(col).lower())]
        df = df.drop(columns=to_drop)

    # Apply the function to each row and filter the DataFrame
    for item in list_to_drop:
        filtered_df = df[~df.apply(contains_extra, axis=1)]

    return filtered_df


#############################################################################################


def fix_merged_headers(df: pd.DataFrame, limit_from_back) -> pd.DataFrame:

    new_cols = {}

    old_name = df.select_dtypes("datetime").columns[0]
    new_name = df.select_dtypes("datetime").columns[0].split("|")[-1]
    df = df.rename(columns={old_name: new_name})

    cols_to_change = df.drop(columns=new_name).columns

    for col in cols_to_change:
        new_cols[col] = "|".join(col.split("|")[limit_from_back:])

    df = df.rename(columns=new_cols)

    return df


#############################################################################################
